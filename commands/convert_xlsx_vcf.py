import os
from openpyxl import load_workbook
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import ContextTypes, ConversationHandler
from commands.vip_system import check_access, send_access_denied, get_user_role, update_user_data, get_user_data
from commands.menu import get_main_menu_keyboard

ASK_FILE, ASK_FILENAME, ASK_CONTACTNAME = range(3)

def create_vcf_from_excel(phone_numbers, contact_name, filename):
    with open(filename, 'w', encoding='utf-8') as f:
        for i, phone in enumerate(phone_numbers, start=1):
            phone_str = str(phone).strip().replace('+', '')
            if phone_str and phone_str.replace('.', '').isnumeric():
                phone_str = phone_str.split('.')[0]
                if not phone_str.startswith('0'):
                    phone_str = '+' + phone_str
                
                vcf_entry = f"""BEGIN:VCARD
VERSION:3.0
FN:{contact_name} {str(i).zfill(4)}
TEL;TYPE=CELL:{phone_str}
END:VCARD

"""
                f.write(vcf_entry)

async def xls_to_vcf_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    
    if not check_access(user_id, "VIP"):
        user_role = get_user_role(user_id)
        await send_access_denied(update, user_role, "VIP")
        return ConversationHandler.END
    
    cancel_keyboard = ReplyKeyboardMarkup([[KeyboardButton("âŒ BATAL âŒ")]], resize_keyboard=True)
    
    text = """```
ğŸš€ XLS TO VCF
â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

Kirim file .xls atau .xlsx yang berisi
nomor telepon untuk dikonversi ke .vcf

â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
```"""
    
    await update.message.reply_text(text, parse_mode="Markdown", reply_markup=cancel_keyboard)
    return ASK_FILE

async def xls_to_vcf_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "âŒ BATAL âŒ":
        keyboard = get_main_menu_keyboard(update.effective_user.id)
        await update.message.reply_text("```\nâŒ Proses dibatalkan\n```",
                parse_mode="Markdown", reply_markup=keyboard)
        return ConversationHandler.END
    
    if not update.message.document:
        await update.message.reply_text("```\nâŒ Kirim file Excel!\n```", parse_mode="Markdown")
        return ASK_FILE
    
    if not (update.message.document.file_name.endswith('.xls') or update.message.document.file_name.endswith('.xlsx')):
        await update.message.reply_text("```\nâŒ File harus berformat .xls atau .xlsx!\n```", parse_mode="Markdown")
        return ASK_FILE
    
    file = await update.message.document.get_file()
    filepath = f"temp_{update.effective_user.id}_{update.message.document.file_name}"
    await file.download_to_drive(filepath)
    
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        all_numbers = []
        
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    all_numbers.append(cell)
        
        phone_numbers = []
        for num in all_numbers:
            num_str = str(num).replace('+', '').strip()
            if num_str and num_str.replace('.', '').replace('-', '').isnumeric():
                phone_numbers.append(num)
        
        if not phone_numbers:
            os.remove(filepath)
            keyboard = get_main_menu_keyboard(update.effective_user.id)
            await update.message.reply_text("```\nâŒ Tidak ada nomor telepon ditemukan!\n```",
                parse_mode="Markdown", reply_markup=keyboard)
            return ConversationHandler.END
        
        context.user_data['phone_numbers'] = phone_numbers
        context.user_data['xls_filepath'] = filepath
        
        cancel_keyboard = ReplyKeyboardMarkup([[KeyboardButton("âŒ BATAL âŒ")]], resize_keyboard=True)
        
        text = f"""```
ğŸ“ NAMA FILE VCF
â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

Total nomor ditemukan: {len(phone_numbers)}

Masukkan nama file output
(tanpa ekstensi .vcf)

Contoh: kontak

â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
```"""
        
        await update.message.reply_text(text, parse_mode="Markdown", reply_markup=cancel_keyboard)
        return ASK_FILENAME
        
    except Exception as e:
        if os.path.exists(filepath):
            os.remove(filepath)
        keyboard = get_main_menu_keyboard(update.effective_user.id)
        await update.message.reply_text(f"```\nâŒ Error reading Excel: {str(e)}\n```",
                parse_mode="Markdown", reply_markup=keyboard)
        return ConversationHandler.END

async def xls_to_vcf_filename(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "âŒ BATAL âŒ":
        if 'xls_filepath' in context.user_data and os.path.exists(context.user_data['xls_filepath']):
            os.remove(context.user_data['xls_filepath'])
        keyboard = get_main_menu_keyboard(update.effective_user.id)
        await update.message.reply_text("```\nâŒ Proses dibatalkan\n```",
                parse_mode="Markdown", reply_markup=keyboard)
        return ConversationHandler.END
    
    context.user_data['vcf_filename'] = update.message.text.strip()
    
    cancel_keyboard = ReplyKeyboardMarkup([[KeyboardButton("âŒ BATAL âŒ")]], resize_keyboard=True)
    
    text = """```
ğŸ‘¤ NAMA KONTAK
â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

Masukkan format nama kontak
(akan ditambah nomor urut otomatis)

Contoh: kontak
Hasil: kontak 0001, kontak 0002, ...

â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
```"""
    
    await update.message.reply_text(text, parse_mode="Markdown", reply_markup=cancel_keyboard)
    return ASK_CONTACTNAME

async def xls_to_vcf_contactname(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "âŒ BATAL âŒ":
        if 'xls_filepath' in context.user_data and os.path.exists(context.user_data['xls_filepath']):
            os.remove(context.user_data['xls_filepath'])
        keyboard = get_main_menu_keyboard(update.effective_user.id)
        await update.message.reply_text("```\nâŒ Proses dibatalkan\n```",
                parse_mode="Markdown", reply_markup=keyboard)
        return ConversationHandler.END
    
    contact_name = update.message.text.strip()
    phone_numbers = context.user_data.get('phone_numbers', [])
    vcf_filename = context.user_data.get('vcf_filename', 'output')
    
    vcf_filepath = f"temp_{update.effective_user.id}_{vcf_filename}.vcf"
    
    create_vcf_from_excel(phone_numbers, contact_name, vcf_filepath)
    
    keyboard = get_main_menu_keyboard(update.effective_user.id)
    
    try:
        await update.message.reply_document(
            document=open(vcf_filepath, 'rb'),
            filename=f"{vcf_filename}.vcf",
            caption=f"âœ… Berhasil convert XLS to VCF!\nğŸ“‚ Total: {len(phone_numbers)} kontak",
            reply_markup=keyboard
        )
        
        user_data = get_user_data(update.effective_user.id)
        total_ops = user_data.get("total_operations", 0) + 1
        update_user_data(update.effective_user.id, {"total_operations": total_ops})
        
    except Exception as e:
        await update.message.reply_text(f"```\nâŒ Error: {str(e)}\n```",
                parse_mode="Markdown", reply_markup=keyboard)
    finally:
        if 'xls_filepath' in context.user_data and os.path.exists(context.user_data['xls_filepath']):
            os.remove(context.user_data['xls_filepath'])
        if os.path.exists(vcf_filepath):
            os.remove(vcf_filepath)
    
    return ConversationHandler.END
